import csv
import io
import logging
import re
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from datetime import datetime
from .validation import validate_drill_score, get_unit_for_drill, DRILL_SCORE_RANGES

logger = logging.getLogger(__name__)

class ImportResult:
    def __init__(self, valid_rows: List[Dict[str, Any]], errors: List[Dict[str, Any]], detected_sport: str = "football", confidence: str = "high", sheets: List[Dict[str, Any]] = None):
        self.valid_rows = valid_rows
        self.errors = errors
        self.detected_sport = detected_sport
        self.confidence = confidence
        self.sheets = sheets or []

class DataImporter:
    """
    Utility class to parse and normalize input data (CSV, Excel, Text)
    into a standardized list of player objects with drill results.
    """
    
    REQUIRED_HEADERS = ['first_name', 'last_name']
    OPTIONAL_HEADERS = ['age_group', 'jersey_number', 'email', 'phone', 'position']
    
    # Map common variations to canonical field names
    FIELD_MAPPING = {
        'first': 'first_name',
        'fname': 'first_name',
        'firstname': 'first_name',
        'last': 'last_name',
        'lname': 'last_name',
        'lastname': 'last_name',
        'jersey': 'jersey_number',
        'number': 'jersey_number',
        'no': 'jersey_number',
        '#': 'jersey_number',
        'age': 'age_group',
        'group': 'age_group',
        'division': 'age_group',
        'pos': 'position'
    }

    @staticmethod
    def _normalize_header(header: str) -> str:
        """Normalize header string to match canonical field names"""
        if not header:
            return ""
        
        clean = str(header).strip().lower().replace(' ', '_').replace('-', '_')
        
        # Check exact matches first
        if clean in DataImporter.FIELD_MAPPING:
            return DataImporter.FIELD_MAPPING[clean]
            
        # Check if it matches any drill keys
        if clean in DRILL_SCORE_RANGES:
            return clean
            
        # Check if it looks like a drill (e.g. "40_yard_dash" -> "40m_dash")
        if '40' in clean and 'dash' in clean:
            return '40m_dash'
        if 'jump' in clean or 'vert' in clean:
            return 'vertical_jump'
        if 'catch' in clean:
            return 'catching'
        if 'throw' in clean:
            return 'throwing'
        if 'agil' in clean or 'l_drill' in clean:
            return 'agility'
            
        return clean

    @staticmethod
    def _clean_value(value: Any) -> Optional[float]:
        """
        Smart Error Correction: Clean and normalize numeric values.
        Handles:
        - Units (s, sec, in, ", etc.)
        - European decimals (4,52 -> 4.52)
        - Typos (4..5 -> 4.5)
        - Whitespace
        """
        if value is None:
            return None
            
        s_val = str(value).strip().lower()
        if not s_val:
            return None
            
        # Remove common units
        s_val = re.sub(r'[a-z"]+$', '', s_val).strip() # Remove trailing units like 's', 'in', '"'
        
        # Replace comma with dot (European decimal)
        s_val = s_val.replace(',', '.')
        
        # Fix double dots (typo)
        s_val = s_val.replace('..', '.')
        
        try:
            return float(s_val)
        except ValueError:
            return None

    @staticmethod
    def _detect_sport(headers: List[str]) -> Tuple[str, str]:
        """
        Auto-detect sport type based on headers.
        Returns (sport, confidence)
        """
        normalized = [DataImporter._normalize_header(h) for h in headers]
        
        # Football indicators
        football_score = 0
        if '40m_dash' in normalized: football_score += 1
        if 'vertical_jump' in normalized: football_score += 1
        if 'agility' in normalized: football_score += 1
        if 'catching' in normalized: football_score += 1
        if 'throwing' in normalized: football_score += 1
        
        if football_score >= 2:
            return 'football', 'high'
        elif football_score == 1:
            return 'football', 'medium'
            
        return 'football', 'low' # Default

    @staticmethod
    def parse_csv(content: bytes) -> ImportResult:
        """Parse CSV content"""
        try:
            # Decode bytes to string
            text = content.decode('utf-8-sig') # Handle BOM if present
            f = io.StringIO(text)
            reader = csv.DictReader(f)
            
            # Normalize headers
            if not reader.fieldnames:
                return ImportResult([], [{"row": 0, "message": "Empty CSV file"}])
                
            normalized_field_map = {
                field: DataImporter._normalize_header(field) 
                for field in reader.fieldnames
            }
            
            sport, confidence = DataImporter._detect_sport(reader.fieldnames)
            
            result = DataImporter._process_rows(reader, normalized_field_map)
            result.detected_sport = sport
            result.confidence = confidence
            return result
            
        except Exception as e:
            logger.error(f"CSV Parse Error: {e}")
            return ImportResult([], [{"row": 0, "message": f"Failed to parse CSV: {str(e)}"}])

    @staticmethod
    def parse_excel(content: bytes, sheet_name: Optional[str] = None) -> ImportResult:
        """
        Parse Excel (XLSX) content.
        If multiple sheets exist and no sheet_name provided, returns list of sheets.
        """
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            
            # Handle multi-sheet detection
            if not sheet_name and len(wb.sheetnames) > 1:
                sheets_info = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    # Get first 3 rows for preview
                    preview = []
                    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
                        preview.append([str(cell or "") for cell in row])
                    sheets_info.append({
                        "name": name,
                        "preview": preview
                    })
                return ImportResult([], [], sheets=sheets_info)
            
            # Select worksheet
            if sheet_name:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    return ImportResult([], [{"row": 0, "message": f"Sheet '{sheet_name}' not found"}])
            else:
                ws = wb.active
            
            # Read rows
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return ImportResult([], [{"row": 0, "message": "Empty Excel sheet"}])
                
            # Extract headers from first row
            headers = [str(cell or "").strip() for cell in rows[0]]
            
            # Map headers
            normalized_field_map = {
                headers[i]: DataImporter._normalize_header(headers[i])
                for i in range(len(headers))
            }
            
            sport, confidence = DataImporter._detect_sport(headers)
            
            # Convert to dict list for processing
            data_rows = []
            for row_idx, row in enumerate(rows[1:], start=2):
                row_data = {}
                has_data = False
                for col_idx, cell_value in enumerate(row):
                    if col_idx < len(headers):
                        key = headers[col_idx]
                        if cell_value is not None and str(cell_value).strip() != "":
                            row_data[key] = cell_value
                            has_data = True
                if has_data:
                    data_rows.append(row_data)
            
            result = DataImporter._process_rows(data_rows, normalized_field_map)
            result.detected_sport = sport
            result.confidence = confidence
            return result
            
        except Exception as e:
            logger.error(f"Excel Parse Error: {e}")
            return ImportResult([], [{"row": 0, "message": f"Failed to parse Excel file: {str(e)}"}])

    @staticmethod
    def parse_image(content: bytes) -> ImportResult:
        """
        Parse image content using OCR.
        """
        try:
            from .ocr import OCRProcessor
            
            lines, confidence = OCRProcessor.extract_rows_from_image(content)
            
            if not lines:
                 return ImportResult([], [{"row": 0, "message": "No text detected in image"}])
                 
            # Convert lines to CSV-like string for parsing
            csv_text = OCRProcessor.lines_to_csv_string(lines)
            
            # Reuse parse_text logic
            result = DataImporter.parse_text(csv_text)
            
            # Override confidence if needed, but for now rely on structure detection
            return result
            
        except ImportError:
            logger.error("OCR dependencies not installed")
            return ImportResult([], [{"row": 0, "message": "OCR system not initialized"}])
        except Exception as e:
            logger.error(f"Image Parse Error: {e}")
            return ImportResult([], [{"row": 0, "message": f"Failed to parse image: {str(e)}"}])

    @staticmethod
    def parse_text(text: str) -> ImportResult:
        """
        Parse pasted text. Assumes either CSV-like structure or specific format.
        For now, implements a robust delimiter sniffer (tab, comma, pipe).
        """
        try:
            # Trim and split lines
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            if not lines:
                return ImportResult([], [{"row": 0, "message": "No text provided"}])
                
            # Sniff delimiter from header row
            header_line = lines[0]
            delimiters = ['\t', ',', '|', ';']
            best_delimiter = ','
            max_cols = 0
            
            for d in delimiters:
                cols = len(header_line.split(d))
                if cols > max_cols:
                    max_cols = cols
                    best_delimiter = d
            
            # Parse using CSV reader with detected delimiter
            f = io.StringIO(text)
            reader = csv.DictReader(f, delimiter=best_delimiter)
            
            if not reader.fieldnames:
                return ImportResult([], [{"row": 0, "message": "Could not parse headers"}])

            normalized_field_map = {
                field: DataImporter._normalize_header(field) 
                for field in reader.fieldnames
            }
            
            sport, confidence = DataImporter._detect_sport(reader.fieldnames)
            
            result = DataImporter._process_rows(reader, normalized_field_map)
            result.detected_sport = sport
            result.confidence = confidence
            return result
            
        except Exception as e:
            logger.error(f"Text Parse Error: {e}")
            return ImportResult([], [{"row": 0, "message": f"Failed to parse text: {str(e)}"}])

    @staticmethod
    def _process_rows(rows: Any, field_map: Dict[str, str]) -> ImportResult:
        """Common processing logic for all input types"""
        valid_rows = []
        errors = []
        
        # Identify which columns map to drill scores
        drill_keys = set(DRILL_SCORE_RANGES.keys())
        
        for idx, row in enumerate(rows, start=1):
            processed_row = {}
            row_errors = []
            
            # Map fields
            for original_key, value in row.items():
                mapped_key = field_map.get(original_key)
                if not mapped_key:
                    continue
                    
                clean_val = str(value).strip() if value is not None else ""
                
                if mapped_key in drill_keys and clean_val:
                    # SMART ERROR CORRECTION: Try to fix common formatting issues
                    cleaned_num = DataImporter._clean_value(clean_val)
                    
                    if cleaned_num is not None:
                        try:
                            validate_drill_score(cleaned_num, mapped_key)
                            processed_row[mapped_key] = cleaned_num
                            
                            # If corrected, maybe we should indicate it?
                            # For now, we just use the corrected value.
                            
                        except Exception as e:
                            row_errors.append(f"Invalid {mapped_key}: {str(e)}")
                            processed_row[f"{mapped_key}_raw"] = clean_val
                    else:
                        row_errors.append(f"Invalid number format for {mapped_key}: '{clean_val}'")
                        processed_row[f"{mapped_key}_raw"] = clean_val
                
                elif mapped_key == 'jersey_number' and clean_val:
                    try:
                        num = int(float(clean_val)) # Handle "10.0" from Excel
                        processed_row[mapped_key] = num
                    except ValueError:
                        row_errors.append(f"Invalid jersey number: {clean_val}")
                        processed_row[f"{mapped_key}_raw"] = clean_val
                
                else:
                    # Regular string fields
                    if clean_val:
                        processed_row[mapped_key] = clean_val

            # Check required fields
            if 'first_name' not in processed_row or not processed_row['first_name']:
                row_errors.append("Missing First Name")
            if 'last_name' not in processed_row or not processed_row['last_name']:
                row_errors.append("Missing Last Name")
                
            # Construct result item
            item = {
                "row_id": idx,
                "data": processed_row,
                "errors": row_errors,
                "original": str(row) # Debug helper
            }
            
            if row_errors:
                errors.append({
                    "row": idx,
                    "data": processed_row,
                    "message": "; ".join(row_errors)
                })
            else:
                valid_rows.append(item)
                
        return ImportResult(valid_rows, errors)
